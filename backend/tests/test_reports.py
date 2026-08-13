"""Reports & Export system (PDF Module 13 / Milestone 4). Drives a debate
session through to completion via the real API, then verifies the reports
list and PDF export both return real generated data (not stubs), and that
access is properly scoped to owner / assigned coach / educator / admin."""
from tests.helpers import create_user_direct


async def _complete_a_debate(client, learner):
    headers = {"Authorization": f"Bearer {learner['access_token']}"}
    start = await client.post(
        "/api/v1/debate/start",
        json={"topic": "Should college be free?", "debate_format": "one_on_one"},
        headers=headers,
    )
    session_id = start.json()["id"]
    await client.post(
        "/api/v1/debate/live",
        json={"session_id": session_id, "text": "Free college increases access to higher education for low-income students."},
        headers=headers,
    )
    await client.post(f"/api/v1/debate/finish?session_id={session_id}", headers=headers)
    return session_id, headers


async def test_reports_list_contains_real_completed_debate(client):
    learner = await create_user_direct("learner", "reports-learner@example.com")
    session_id, headers = await _complete_a_debate(client, learner)

    res = await client.get("/api/v1/reports", headers=headers)
    assert res.status_code == 200
    body = res.json()
    assert body["total"] >= 1
    match = next(item for item in body["items"] if item["id"] == session_id)
    assert match["topic"] == "Should college be free?"
    assert match["overall_score"] is not None  # a real computed score, not null/placeholder


async def test_pdf_export_returns_real_pdf_bytes(client):
    learner = await create_user_direct("learner", "reports-pdf-learner@example.com")
    session_id, headers = await _complete_a_debate(client, learner)

    res = await client.get(f"/api/v1/reports/{session_id}/pdf", headers=headers)
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/pdf"
    # A real generated PDF, not an empty/placeholder file — check the PDF
    # magic bytes and a non-trivial size.
    assert res.content[:4] == b"%PDF"
    assert len(res.content) > 1000


async def test_single_session_excel_export_returns_real_xlsx_bytes(client):
    learner = await create_user_direct("learner", "reports-xlsx-learner@example.com")
    session_id, headers = await _complete_a_debate(client, learner)

    res = await client.get(f"/api/v1/reports/{session_id}/excel", headers=headers)
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # .xlsx is a zip container — "PK" magic bytes confirm a real workbook
    # was generated, not an empty/placeholder file.
    assert res.content[:2] == b"PK"
    assert len(res.content) > 1000

    # Parse it back and confirm it actually contains the real topic —
    # not just well-formed bytes with no real data inside.
    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(res.content))
    assert "Summary" in wb.sheetnames
    summary_values = [cell.value for row in wb["Summary"].iter_rows() for cell in row]
    assert "Should college be free?" in summary_values


async def test_bulk_reports_excel_export_reflects_real_sessions(client):
    learner = await create_user_direct("learner", "reports-xlsx-bulk@example.com")
    await _complete_a_debate(client, learner)
    _, headers = await _complete_a_debate(client, learner)

    res = await client.get("/api/v1/reports/excel", headers=headers)
    assert res.status_code == 200
    assert res.content[:2] == b"PK"

    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(res.content))
    ws = wb["Debate Reports"]
    # Header row + at least 2 real completed-debate rows (not fabricated).
    assert ws.max_row >= 3


async def test_bulk_excel_export_respects_report_access_control(client):
    """Reuses list_my_reports()'s own authorization rule — confirm an
    unrelated learner's bulk export never contains another learner's data."""
    owner = await create_user_direct("learner", "reports-xlsx-owner@example.com")
    await _complete_a_debate(client, owner)

    stranger = await create_user_direct("learner", "reports-xlsx-stranger@example.com")
    res = await client.get("/api/v1/reports/excel", headers=await _headers_for(stranger))
    assert res.status_code == 200

    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(res.content))
    ws = wb["Debate Reports"]
    assert ws.max_row == 1  # header only, no rows — stranger has zero completed debates


async def _headers_for(user):
    return {"Authorization": f"Bearer {user['access_token']}"}


async def test_reports_are_not_visible_to_unrelated_learner(client):
    learner = await create_user_direct("learner", "reports-owner2@example.com")
    session_id, _ = await _complete_a_debate(client, learner)

    stranger = await create_user_direct("learner", "reports-stranger@example.com")
    stranger_headers = {"Authorization": f"Bearer {stranger['access_token']}"}

    # A learner's learner_id query param is ignored server-side — they
    # always see only their own reports, regardless of what id they pass.
    # That's safe (no cross-user data is returned) even though the status
    # code is 200 rather than a 403: confirm no leak rather than assume a
    # specific status code.
    res = await client.get(f"/api/v1/reports?learner_id={learner['id']}", headers=stranger_headers)
    assert res.status_code == 200
    assert res.json()["total"] == 0
    assert res.json()["items"] == []

    pdf_res = await client.get(f"/api/v1/reports/{session_id}/pdf", headers=stranger_headers)
    assert pdf_res.status_code == 404


async def test_assigned_coach_can_view_but_unassigned_coach_cannot(client):
    learner = await create_user_direct("learner", "reports-coached-learner@example.com")
    session_id, _ = await _complete_a_debate(client, learner)

    from app.core import database as db

    assigned_coach = await create_user_direct("debate_coach", "assigned-coach@example.com")
    await db.coach_assignments_collection.insert_one(
        {"coach_id": assigned_coach["id"], "learner_id": learner["id"]}
    )
    unassigned_coach = await create_user_direct("debate_coach", "unassigned-coach@example.com")

    ok = await client.get(
        f"/api/v1/reports?learner_id={learner['id']}",
        headers={"Authorization": f"Bearer {assigned_coach['access_token']}"},
    )
    assert ok.status_code == 200
    assert ok.json()["total"] >= 1

    denied = await client.get(
        f"/api/v1/reports?learner_id={learner['id']}",
        headers={"Authorization": f"Bearer {unassigned_coach['access_token']}"},
    )
    assert denied.status_code == 403

    denied_pdf = await client.get(
        f"/api/v1/reports/{session_id}/pdf",
        headers={"Authorization": f"Bearer {unassigned_coach['access_token']}"},
    )
    assert denied_pdf.status_code == 404
