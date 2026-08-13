"""
Reports Module (Milestone 4) — real Excel (.xlsx) generation.

Mirrors report_pdf_service.py's approach exactly: built purely from data
already persisted for a completed debate — no template/filler content, no
fabricated rows. Two exports are provided, matching the two natural shapes
"Excel export" takes in the PDF:
  - build_debate_report_xlsx()  — one workbook, single debate session,
    same level of detail as the PDF report (mirrors it, doesn't duplicate
    its data-gathering — the router fetches once and passes to both).
  - build_reports_summary_xlsx() — one workbook, every completed session a
    user is authorized to see, one row per session — for bulk analysis in
    a spreadsheet (sorting/filtering/charting outside the app).
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header_row(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _autosize_columns(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_debate_report_xlsx(
    *,
    topic: str,
    debate_format: str,
    date: str,
    overall_score: float | None,
    report: dict | None,
    fallacies: list[dict],
    coach_score: float | None = None,
    coach_comments: str | None = None,
    educator_score: float | None = None,
    educator_comments: str | None = None,
    session_id: str | None = None,
    learner_name: str | None = None,
    learner_email: str | None = None,
    presentation: dict | None = None,
) -> bytes:
    wb = Workbook()

    # --- Sheet 1: Summary ---------------------------------------------------
    summary = wb.active
    summary.title = "Summary"
    rows = [
        ("Topic", topic),
        ("Debate Format", debate_format.replace("_", " ").title()),
        ("Date", date),
        ("Learner", learner_name or "—"),
        ("Learner Email", learner_email or "—"),
        ("Overall Score", overall_score if overall_score is not None else "Not yet scored"),
    ]
    if report:
        rows += [
            ("Argument Quality", report.get("argument_quality", "—")),
            ("Evidence Usage", report.get("evidence_usage", "—")),
            ("Logical Consistency", report.get("logical_consistency", "—")),
            ("Rebuttal Effectiveness", report.get("rebuttal_effectiveness", "—")),
            ("Communication Skills", report.get("communication_skills", "—")),
        ]
    if coach_score is not None:
        rows.append(("Coach Score", coach_score))
    if coach_comments:
        rows.append(("Coach Comments", coach_comments))
    if educator_score is not None:
        rows.append(("Educator Score", educator_score))
    if educator_comments:
        rows.append(("Educator Comments", educator_comments))
    if presentation and presentation.get("speech_metrics"):
        metrics = presentation["speech_metrics"]
        rows += [
            ("Speech Pace (WPM)", metrics.get("words_per_minute", "—")),
            ("Filler Word Count", metrics.get("filler_word_count", "—")),
            ("Confidence Score", metrics.get("confidence_score", "—")),
            ("Clarity Score", metrics.get("clarity_score", "—")),
        ]

    summary.append(["Field", "Value"])
    _style_header_row(summary, 1, 2)
    for label, value in rows:
        summary.append([label, value])
    _autosize_columns(summary, {1: 24, 2: 50})

    # --- Sheet 2: Feedback (only if a report exists) ------------------------
    if report:
        fb = wb.create_sheet("Feedback")
        fb.append(["Category", "Item"])
        _style_header_row(fb, 1, 2)
        for category, key in (
            ("Strength", "strengths"),
            ("Weakness", "weaknesses"),
            ("Missing Evidence", "missing_evidence"),
            ("Logical Issue", "logical_issues"),
            ("Recommended Improvement", "recommended_improvements"),
            ("Learning Recommendation", "learning_recommendations"),
        ):
            for item in report.get(key, []) or []:
                fb.append([category, item])
        _autosize_columns(fb, {1: 24, 2: 70})
        if report.get("final_summary"):
            fb.append([])
            fb.append(["Final Summary", report["final_summary"]])

    # --- Sheet 3: Fallacies (only if any were detected) ---------------------
    if fallacies:
        fl = wb.create_sheet("Fallacies Detected")
        fl.append(["Type", "Severity", "Excerpt", "Explanation", "Suggested Correction"])
        _style_header_row(fl, 1, 5)
        for f in fallacies:
            fl.append([
                f.get("fallacy_type", "—"),
                f.get("severity", "—"),
                f.get("offending_text", "—"),
                f.get("explanation", "—"),
                f.get("correction_suggestion", "—"),
            ])
        _autosize_columns(fl, {1: 18, 2: 12, 3: 40, 4: 50, 5: 50})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_reports_summary_xlsx(items: list[dict]) -> bytes:
    """One row per completed debate session — same fields as the /reports
    list endpoint's JSON, for bulk spreadsheet analysis. `items` is exactly
    what reports.list_my_reports() already assembles from real data; this
    function does no additional DB access and fabricates nothing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Debate Reports"

    headers = [
        "Topic", "Debate Format", "Date", "Overall Score",
        "Coach Score", "Coach Comments", "Educator Score", "Educator Comments", "Review Status",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for item in items:
        ws.append([
            item.get("topic", "—"),
            (item.get("debate_format") or "—").replace("_", " ").title(),
            item.get("date", "—"),
            item.get("overall_score") if item.get("overall_score") is not None else "Not yet scored",
            item.get("coach_score") if item.get("coach_score") is not None else "—",
            item.get("coach_comments") or "—",
            item.get("educator_score") if item.get("educator_score") is not None else "—",
            item.get("educator_comments") or "—",
            item.get("review_status") or "—",
        ])

    _autosize_columns(ws, {1: 40, 2: 18, 3: 22, 4: 14, 5: 12, 6: 40, 7: 14, 8: 40, 9: 16})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
